from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from datetime import datetime
import csv
import os
from typing import List, Dict

def export_to_excel(transactions: List[Dict], filename: str, summary_stats: Dict = None) -> str:
    """Export transactions to Excel file with formatting"""
    try:
        wb = Workbook()
        
        # Transactions sheet
        ws = wb.active
        ws.title = "Transactions"
        
        # Headers
        headers = ['Date', 'Type', 'Category', 'Amount (₹)', 'Reason', 'Balance (₹)']
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add data
        for trans in transactions:
            date_str = datetime.fromisoformat(trans['transaction_date']).strftime('%d %b %Y, %I:%M %p')
            ws.append([
                date_str,
                trans['type'].capitalize(),
                trans['category'],
                trans['amount'],
                trans['reason'] or '',
                trans['balance_after']
            ])
        
        # Style data rows
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Color code by type
            trans_type = row[1].value
            if trans_type == 'Credit':
                row[1].fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                row[1].font = Font(color="155724", bold=True)
            else:
                row[1].fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                row[1].font = Font(color="721C24", bold=True)
            
            # Format amount columns
            row[3].number_format = '#,##0.00'
            row[5].number_format = '#,##0.00'
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 40
        ws.column_dimensions['F'].width = 15
        
        # Summary sheet
        if summary_stats:
            ws_summary = wb.create_sheet("Summary")
            ws_summary.append(['Payment Tracker Summary Report'])
            ws_summary.append(['Generated on:', datetime.now().strftime('%d %B %Y, %I:%M %p')])
            ws_summary.append([])
            
            ws_summary.append(['Metric', 'Value (₹)'])
            ws_summary.append(['Current Balance', summary_stats.get('current_balance', 0)])
            ws_summary.append(['Total Credit', summary_stats.get('total_credit', 0)])
            ws_summary.append(['Total Debit', summary_stats.get('total_debit', 0)])
            ws_summary.append(['Today\'s Expense', summary_stats.get('today_expense', 0)])
            
            # Style summary
            ws_summary['A1'].font = Font(size=16, bold=True)
            ws_summary.merge_cells('A1:B1')
            
            for row in ws_summary.iter_rows(min_row=4, max_row=8):
                row[0].font = Font(bold=True)
                row[1].number_format = '#,##0.00'
            
            ws_summary.column_dimensions['A'].width = 20
            ws_summary.column_dimensions['B'].width = 20
        
        # Save file
        filepath = os.path.join('exports', filename)
        wb.save(filepath)
        
        return filepath
    except Exception as e:
        raise Exception(f"Excel export failed: {str(e)}")

def export_to_pdf(transactions: List[Dict], filename: str, summary_stats: Dict = None) -> str:
    """Export transactions to PDF file"""
    try:
        filepath = os.path.join('exports', filename)
        doc = SimpleDocTemplate(filepath, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#366092'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        elements.append(Paragraph("Payment Tracker Report", title_style))
        
        # Date
        date_style = ParagraphStyle(
            'DateStyle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.grey,
            alignment=TA_CENTER,
            spaceAfter=20
        )
        elements.append(Paragraph(f"Generated on: {datetime.now().strftime('%d %B %Y, %I:%M %p')}", date_style))
        
        # Summary section
        if summary_stats:
            elements.append(Paragraph("Summary", styles['Heading2']))
            
            summary_data = [
                ['Metric', 'Amount (₹)'],
                ['Current Balance', f"₹ {summary_stats.get('current_balance', 0):,.2f}"],
                ['Total Credit', f"₹ {summary_stats.get('total_credit', 0):,.2f}"],
                ['Total Debit', f"₹ {summary_stats.get('total_debit', 0):,.2f}"],
                ['Today\'s Expense', f"₹ {summary_stats.get('today_expense', 0):,.2f}"]
            ]
            
            summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
            ]))
            
            elements.append(summary_table)
            elements.append(Spacer(1, 20))
        
        # Transactions section
        elements.append(Paragraph("Transaction History", styles['Heading2']))
        elements.append(Spacer(1, 10))
        
        # Prepare transaction data
        trans_data = [['Date', 'Type', 'Category', 'Amount (₹)', 'Balance (₹)']]
        
        for trans in transactions[:100]:  # Limit to 100 transactions for PDF
            date_str = datetime.fromisoformat(trans['transaction_date']).strftime('%d %b %Y')
            trans_data.append([
                date_str,
                trans['type'].capitalize(),
                trans['category'],
                f"₹ {trans['amount']:,.2f}",
                f"₹ {trans['balance_after']:,.2f}"
            ])
        
        # Create table
        trans_table = Table(trans_data, colWidths=[1.5*inch, 0.8*inch, 1.2*inch, 1.2*inch, 1.3*inch])
        trans_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (3, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        
        elements.append(trans_table)
        
        if len(transactions) > 100:
            elements.append(Spacer(1, 10))
            elements.append(Paragraph(f"Note: Showing first 100 of {len(transactions)} transactions", 
                                     styles['Normal']))
        
        # Build PDF
        doc.build(elements)
        
        return filepath
    except Exception as e:
        raise Exception(f"PDF export failed: {str(e)}")

def export_to_csv(transactions: List[Dict], filename: str) -> str:
    """Export transactions to CSV file"""
    try:
        filepath = os.path.join('exports', filename)
        
        with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
            fieldnames = ['Date', 'Type', 'Category', 'Amount', 'Reason', 'Balance']
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            
            writer.writeheader()
            
            for trans in transactions:
                date_str = datetime.fromisoformat(trans['transaction_date']).strftime('%d %b %Y, %I:%M %p')
                writer.writerow({
                    'Date': date_str,
                    'Type': trans['type'].capitalize(),
                    'Category': trans['category'],
                    'Amount': trans['amount'],
                    'Reason': trans['reason'] or '',
                    'Balance': trans['balance_after']
                })
        
        return filepath
    except Exception as e:
        raise Exception(f"CSV export failed: {str(e)}")